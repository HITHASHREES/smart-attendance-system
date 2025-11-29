import tkinter as tk
from tkinter import ttk, messagebox
import pandas as pd
import os
import gspread
import threading
import subprocess
from datetime import datetime, timedelta
from google.oauth2.service_account import Credentials

# ⬅️ Student registration window
from register_student import open_registration_window

# ======================================================
#  SUBJECTS (Multi-subject support)
# ======================================================
SUBJECTS = ["DBMS", "Java", "Python", "OS", "CN"]


# ======================================================
#  TOOLTIP CLASS (for hover descriptions)
# ======================================================
class ToolTip:
    def __init__(self, widget, text, delay=600):
        self.widget = widget
        self.text = text
        self.delay = delay
        self._id = None
        self.tipwindow = None

        widget.bind("<Enter>", self._schedule)
        widget.bind("<Leave>", self._hide)
        widget.bind("<ButtonPress>", self._hide)

    def _schedule(self, event=None):
        self._unschedule()
        self._id = self.widget.after(self.delay, self._show)

    def _unschedule(self):
        if self._id is not None:
            self.widget.after_cancel(self._id)
            self._id = None

    def _show(self, event=None):
        if self.tipwindow or not self.text:
            return

        # position near mouse pointer
        x = self.widget.winfo_rootx() + 40
        y = self.widget.winfo_rooty() + 25

        self.tipwindow = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)        # borderless
        tw.wm_geometry(f"+{x}+{y}")

        frame = tk.Frame(
            tw,
            bg="#020617",
            padx=6,
            pady=4,
            highlightthickness=1,
            highlightbackground="#38BDF8"
        )
        frame.pack()

        label = tk.Label(
            frame,
            text=self.text,
            justify="left",
            bg="#020617",
            fg="#E5E7EB",
            font=("Segoe UI", 9),
            wraplength=260
        )
        label.pack()

    def _hide(self, event=None):
        self._unschedule()
        if self.tipwindow is not None:
            self.tipwindow.destroy()
            self.tipwindow = None


# -----------------------------
# 📁 File paths
# -----------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# NOTE: this file is still used for some global operations
ATTENDANCE_FILE = os.path.join(BASE_DIR, "attendance.csv")

# NOTE: use the same JSON as login.py
SERVICE_ACCOUNT_FILE = os.path.join(BASE_DIR, "service_account.json")
STUDENTS_FILE = os.path.join(BASE_DIR, "students.csv")
BACKUP_DIR = os.path.join(BASE_DIR, "backups")
ICONS_DIR = os.path.join(BASE_DIR, "icons")  # folder with sidebar icons

os.makedirs(BACKUP_DIR, exist_ok=True)


# -----------------------------
# 📦 Backup helper
# -----------------------------
def create_backup():
    if not os.path.exists(ATTENDANCE_FILE):
        return None
    try:
        ts = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        backup_path = os.path.join(BACKUP_DIR, f"attendance_backup_{ts}.csv")
        pd.read_csv(ATTENDANCE_FILE).to_csv(backup_path, index=False)
        print(f"📦 Backup created at: {backup_path}")
        return backup_path
    except Exception as e:
        print("Backup failed:", e)
        return None


# -----------------------------
# 🆕 Mark attendance function
# (still simple, used if some scripts write to attendance.csv)
# -----------------------------
def mark_attendance(usn, name):
    today = datetime.now().strftime("%Y-%m-%d")
    time_now = datetime.now().strftime("%H:%M:%S")

    # create file if not exists
    if not os.path.exists(ATTENDANCE_FILE):
        with open(ATTENDANCE_FILE, "w", encoding="utf-8") as f:
            f.write("USN,Name,Date,Time,Status\n")

    already = False
    with open(ATTENDANCE_FILE, "r", encoding="utf-8") as f:
        for row in f:
            row = row.strip().split(",")
            if len(row) >= 3 and row[0] == usn and row[2] == today:
                already = True
                break

    if not already:
        with open(ATTENDANCE_FILE, "a", encoding="utf-8") as f:
            f.write(f"{usn},{name},{today},{time_now},PRESENT\n")

    print(f"✔ Attendance marked → {usn} - {name}")


# -----------------------------
# ☁ Google Sheets Setup
# -----------------------------
SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]


def get_google_sheet_data():
    try:
        creds = Credentials.from_service_account_file(
            SERVICE_ACCOUNT_FILE, scopes=SCOPES
        )
        client = gspread.authorize(creds)
        sheet = client.open("Attendance Report").sheet1
        return pd.DataFrame(sheet.get_all_records())
    except Exception as e:
        messagebox.showerror("Error", f"Could not load Google Sheet:\n{e}")
        return None


# -----------------------------
# 📄 Load Local CSV (global one)
# -----------------------------
def get_local_data():
    if not os.path.exists(ATTENDANCE_FILE):
        return None
    try:
        return pd.read_csv(ATTENDANCE_FILE)
    except Exception:
        return None


# -----------------------------
# 🧹 Clean Data
# -----------------------------
def clean_data(df):
    try:
        df.columns = df.columns.str.strip()
        return df.drop_duplicates(subset=["USN", "Date"], keep="last")
    except Exception:
        return df


# -----------------------------
# 📊 UI Class
# -----------------------------
class AttendanceDashboard:
    def __init__(self, root):
        self.root = root
        self.root.title("Smart Attendance - Admin Dashboard")
        self.root.geometry("1300x750")

        # Midnight theme
        self.root.config(bg="#020617")  # deep black-blue

        self.current_date = datetime.now().strftime("%Y-%m-%d")

        # currently selected subject (for face recognition, absent view)
        self.current_subject = tk.StringVar(value=SUBJECTS[0])

        # keep references to PhotoImage so they don't get garbage collected
        self.icons = {}
        self._load_all_icons()

        # sidebar indicator animation
        self.sidebar_indicator = None
        self.indicator_y = 0

        # =========================
        # MAIN LAYOUT: sidebar + main
        # =========================
        self.sidebar = tk.Frame(
            root,
            bg="#020617",  # pure midnight
            width=220,
            highlightthickness=1,
            highlightbackground="#111827",
        )
        self.sidebar.pack(side="left", fill="y")

        self.main = tk.Frame(
            root,
            bg="#030712",  # slightly lighter dark
            highlightthickness=0,
        )
        self.main.pack(side="right", fill="both", expand=True)

        # -------------------------
        # Sidebar content
        # -------------------------
        logo = tk.Label(
            self.sidebar,
            text="ADMIN\nPANEL",
            fg="#E5E7EB",
            bg="#020617",
            font=("Segoe UI", 18, "bold"),
            pady=20,
        )
        logo.pack()

        self.sidebar_buttons = []
        self.active_sidebar = None

        # sidebar buttons + tooltips
        self._add_sidebar_button(
            "Dashboard",
            self.show_dashboard_view,
            "dashboard",
            tooltip="Overview cards and today's attendance table",
        )
        self._add_sidebar_button(
            "Attendance",
            self.show_today_attendance,
            "attendance",
            tooltip="View attendance records (today view by default)",
        )
        self._add_sidebar_button(
            "Students",
            self.open_student_registration,
            "students",
            tooltip="Add or manage registered students",
        )
        self._add_sidebar_button(
            "Face Recognition",
            self.run_face_recognition_thread,
            "face",
            tooltip="Open camera and mark attendance using face recognition",
        )
        self._add_sidebar_button(
            "Export",
            self.open_export_menu,
            "export",
            tooltip="Export today's, previous day's or summary attendance",
        )
        self._add_sidebar_button(
            "Visualization",
            self.open_visualization_menu,
            "chart",
            tooltip="View charts for class overview or individual trends",
        )

        tk.Label(self.sidebar, text="", bg="#020617").pack(expand=True, fill="y")

        # Logout button with icon
        logout_btn = tk.Button(
            self.sidebar,
            text="  Logout / Exit",
            font=("Segoe UI", 11, "bold"),
            bg="#DC2626",
            fg="white",
            bd=0,
            relief="flat",
            command=self.root.quit,
            padx=10,
            pady=8,
            image=self.icons.get("logout"),
            compound="left",
            anchor="w",
            activebackground="#B91C1C",
            activeforeground="white",
        )
        logout_btn.pack(fill="x", padx=20, pady=20)
        self._add_hover_effect(logout_btn, "#B91C1C", "#DC2626")
        ToolTip(logout_btn, "Close the admin dashboard window")

        # -------------------------
        # Top bar in main area
        # -------------------------
        top_bar = tk.Frame(
            self.main,
            bg="#020617",
            height=60,
            highlightthickness=1,
            highlightbackground="#111827",
        )
        top_bar.pack(fill="x", padx=10, pady=10)

        self.title_label = tk.Label(
            top_bar,
            text="Smart Attendance Dashboard (Admin)",
            font=("Segoe UI", 18, "bold"),
            bg="#020617",
            fg="#E5E7EB",
        )
        self.title_label.pack(side="left", padx=20)

        # Subject chooser (right side)
        subject_frame = tk.Frame(top_bar, bg="#020617")
        subject_frame.pack(side="right", padx=20)

        tk.Label(
            subject_frame,
            text="Subject:",
            font=("Segoe UI", 10, "bold"),
            bg="#020617",
            fg="#9CA3AF",
        ).grid(row=0, column=0, padx=(0, 6))

        subject_combo = ttk.Combobox(
            subject_frame,
            textvariable=self.current_subject,
            values=SUBJECTS,
            state="readonly",
            width=12,
        )
        subject_combo.grid(row=0, column=1)
        subject_combo.current(0)
        ToolTip(subject_combo, "Select the subject for which you are taking attendance")

        self.date_label = tk.Label(
            top_bar,
            text=datetime.now().strftime("%A, %d %b %Y"),
            font=("Segoe UI", 11),
            bg="#020617",
            fg="#9CA3AF",
        )
        self.date_label.pack(side="right", padx=20)

        # -------------------------
        # Overview cards
        # -------------------------
        cards_frame = tk.Frame(self.main, bg="#030712")
        cards_frame.pack(fill="x", padx=20, pady=(0, 10))

        self.card_total_students = self._create_stat_card(
            cards_frame, "Total Students", "0", "#38BDF8"
        )
        self.card_present_today = self._create_stat_card(
            cards_frame, "Present Today", "0", "#34D399"
        )
        self.card_absent_today = self._create_stat_card(
            cards_frame, "Absent Today", "0", "#F97373"
        )
        self.card_total_classes = self._create_stat_card(
            cards_frame, "Total Classes", "0", "#A855F7"
        )

        # -------------------------
        # Action buttons panel
        # -------------------------
        actions_frame = tk.LabelFrame(
            self.main,
            text="Quick Actions",
            font=("Segoe UI", 11, "bold"),
            bg="#030712",
            fg="#E5E7EB",
            padx=10,
            pady=10,
            highlightthickness=1,
            highlightbackground="#111827",
        )
        actions_frame.pack(fill="x", padx=20, pady=5)

        def make_action_btn(row, col, text, cmd, bg, tooltip_text=None):
            btn = tk.Button(
                actions_frame,
                text=text,
                command=cmd,
                bg=bg,
                fg="white",
                font=("Segoe UI", 10, "bold"),
                padx=10,
                pady=5,
                relief="flat",
                bd=0,
                activebackground="#111827",
                activeforeground="#E5E7EB",
            )
            btn.grid(row=row, column=col, padx=5, pady=5, sticky="ew")
            self._add_hover_effect(btn, self._slightly_darker(bg), bg)
            if tooltip_text:
                ToolTip(btn, tooltip_text)

        # row 0
        make_action_btn(
            0, 0,
            "🔄 Refresh Sheets",
            self.load_google_data,
            "#2563EB",
            "Reload attendance directly from the Google Sheet"
        )
        make_action_btn(
            0, 1,
            "📂 Load Local CSV",
            self.load_local_data,
            "#4B5563",
            "Load attendance data from the local CSV file"
        )
        make_action_btn(
            0, 2,
            "🗓 Show Today",
            self.show_today_attendance,
            "#F97316",
            "Filter the table to show only today's attendance"
        )
        make_action_btn(
            0, 3,
            "📸 Start Face Recognition",
            self.run_face_recognition_thread,
            "#059669",
            "Open the camera and mark students present using face recognition"
        )

        # row 1
        make_action_btn(
            1, 0,
            "📥 Export Today",
            self.export_todays_attendance,
            "#0EA5E9",
            "Export today's present and absent students to Excel"
        )
        make_action_btn(
            1, 1,
            "📦 Export Previous Day",
            self.export_previous_day_attendance,
            "#6B21A8",
            "Export attendance report for yesterday"
        )
        make_action_btn(
            1, 2,
            "📤 Export Summary",
            self.export_summary_to_excel,
            "#1D4ED8",
            "Export full attendance summary with percentages"
        )
        make_action_btn(
            1, 3,
            "➕ Add New Student",
            self.open_student_registration,
            "#22C55E",
            "Register a new student with face images"
        )

        # row 2
        make_action_btn(
            2, 0,
            "🧽 Clear Attendance",
            self.clear_previous_attendance,
            "#DC2626",
            "Clear all attendance records (after creating a backup)"
        )
        make_action_btn(
            2, 1,
            "💾 Backup Now",
            self.manual_backup,
            "#4B5563",
            "Create a manual backup of the current attendance CSV"
        )
        make_action_btn(
            2, 2,
            "📊 Class Overview",
            self.open_class_overview_window,
            "#0F766E",
            "View graphs of class-wise attendance over time"
        )
        make_action_btn(
            2, 3,
            "📈 Student Trend (Select Row)",
            self._show_student_trend_internal,
            "#7C2D12",
            "After selecting a student, see their attendance trend chart"
        )

        # row 3 – Search + Absent buttons
        make_action_btn(
            3, 0,
            "🔍 Search Student",
            self.open_search_student_window,
            "#6366F1",
            "Search a particular student's attendance by USN / Name"
        )
        make_action_btn(
            3, 1,
            "🚫 View Absent (Subject)",
            self.show_absent_subjectwise,
            "#DC2626",
            "Show students absent for the selected subject today"
        )

        for i in range(4):
            actions_frame.columnconfigure(i, weight=1)

        # -------------------------
        # Table
        # -------------------------
        table_container = tk.Frame(self.main, bg="#030712")
        table_container.pack(fill="both", expand=True, padx=20, pady=10)

        self.tree = ttk.Treeview(
            table_container,
            columns=("USN", "Name", "Date", "Time", "Status"),
            show="headings",
            height=15,
        )

        style = ttk.Style()
        style.theme_use("default")
        style.configure(
            "Treeview",
            font=("Segoe UI", 10),
            rowheight=24,
            background="#020617",
            fieldbackground="#020617",
            foreground="#E5E7EB",
        )
        style.configure(
            "Treeview.Heading",
            font=("Segoe UI", 10, "bold"),
            background="#020617",
            foreground="#E5E7EB",
        )
        style.map(
            "Treeview",
            background=[("selected", "#1D4ED8")],
            foreground=[("selected", "#E5E7EB")],
        )

        # striped rows
        self.tree.tag_configure("oddrow", background="#030712")
        self.tree.tag_configure("evenrow", background="#020617")

        for col in ("USN", "Name", "Date", "Time", "Status"):
            self.tree.heading(col, text=col)
            self.tree.column(col, width=170, anchor="center")

        vsb = ttk.Scrollbar(table_container, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscroll=vsb.set)
        vsb.pack(side="right", fill="y")
        self.tree.pack(fill="both", expand=True)

        self.tree.bind("<Double-1>", self.on_tree_double_click)

        # Start background tasks
        self.schedule_midnight_check()
        self.refresh_overview_cards()
        self.show_dashboard_view()  # default view

    # =========================
    # Icon helpers
    # =========================
    def _load_icon(self, filename):
        path = os.path.join(ICONS_DIR, filename)
        if not os.path.exists(path):
            return None
        try:
            img = tk.PhotoImage(file=path)

            # 🔽 Smart downscale big icons (like 768x768) to around 32px
            max_size = 32
            factor_w = max(1, img.width() // max_size)
            factor_h = max(1, img.height() // max_size)
            factor = max(factor_w, factor_h)
            if factor > 1:
                img = img.subsample(factor, factor)

            return img
        except Exception as e:
            print("Icon load failed:", filename, e)
            return None

    def _load_all_icons(self):
        self.icons["dashboard"] = self._load_icon("dashboard.png")
        self.icons["attendance"] = self._load_icon("attendance.png")
        self.icons["students"] = self._load_icon("students.png")
        self.icons["face"] = self._load_icon("face.png")
        self.icons["export"] = self._load_icon("export.png")
        self.icons["chart"] = self._load_icon("chart.png")
        self.icons["logout"] = self._load_icon("logout.png")

    # =========================
    # Sidebar / hover helpers
    # =========================
    def _add_hover_effect(self, widget, hover_color, normal_color):
        """Hover glow + bg change for buttons."""
        def on_enter(e):
            widget.config(
                bg=hover_color,
                highlightthickness=2,
                highlightbackground="#38BDF8",
                highlightcolor="#38BDF8",
            )

        def on_leave(e):
            widget.config(
                bg=normal_color,
                highlightthickness=0
            )

        widget.bind("<Enter>", on_enter)
        widget.bind("<Leave>", on_leave)

    def _slightly_darker(self, hex_color):
        # tiny darkening for hover effect
        hex_color = hex_color.lstrip("#")
        r = max(0, int(hex_color[0:2], 16) - 15)
        g = max(0, int(hex_color[2:4], 16) - 15)
        b = max(0, int(hex_color[4:6], 16) - 15)
        return f"#{r:02x}{g:02x}{b:02x}"

    def _add_sidebar_button(self, text, command, icon_key=None, tooltip=None):
        icon = self.icons.get(icon_key)
        btn = tk.Button(
            self.sidebar,
            text=f"  {text}",
            font=("Segoe UI", 11),
            bg="#020617",
            fg="#9CA3AF",
            bd=0,
            relief="flat",
            anchor="w",
            padx=18,
            pady=8,
            activebackground="#111827",
            activeforeground="#E5E7EB",
            image=icon,
            compound="left" if icon is not None else None,
        )

        # command added after creation so we can pass btn into lambda
        btn.config(
            command=lambda n=text, c=command, b=btn: self._on_sidebar_click(n, c, b)
        )

        btn.pack(fill="x", pady=2)
        self._add_hover_effect(btn, "#111827", "#020617")
        if tooltip:
            ToolTip(btn, tooltip)
        self.sidebar_buttons.append(btn)

        # create indicator on the first button
        if len(self.sidebar_buttons) == 1:
            self.sidebar.update_idletasks()
            self._init_sidebar_indicator(btn)

    def _init_sidebar_indicator(self, btn):
        y = btn.winfo_y()
        h = btn.winfo_height()
        self.indicator_y = y
        self.sidebar_indicator = tk.Frame(
            self.sidebar,
            bg="#38BDF8",
            width=3,
            height=max(10, h - 4),
        )
        self.sidebar_indicator.place(x=2, y=self.indicator_y + 2)

    def _animate_indicator_to(self, btn):
        if self.sidebar_indicator is None:
            return

        self.sidebar.update_idletasks()
        target_y = btn.winfo_y()
        h = btn.winfo_height()
        self.sidebar_indicator.config(height=max(10, h - 4))

        step = 4

        def slide():
            cur = self.indicator_y
            if abs(cur - target_y) <= step:
                self.indicator_y = target_y
                self.sidebar_indicator.place_configure(y=self.indicator_y + 2)
                return
            if target_y > cur:
                self.indicator_y = cur + step
            else:
                self.indicator_y = cur - step
            self.sidebar_indicator.place_configure(y=self.indicator_y + 2)
            self.root.after(10, slide)

        slide()

    def _on_sidebar_click(self, name, cmd, btn):
        # highlight active & animate indicator
        for b in self.sidebar_buttons:
            if b is btn:
                b.config(
                    bg="#1D4ED8",
                    fg="#E5E7EB",
                    highlightthickness=2,
                    highlightbackground="#38BDF8",
                )
            else:
                b.config(bg="#020617", fg="#9CA3AF", highlightthickness=0)

        self._animate_indicator_to(btn)
        self.active_sidebar = name
        cmd()

    def show_dashboard_view(self):
        """Default view = today + stats."""
        self.show_today_attendance()
        self.refresh_overview_cards()

    # =========================
    # Overview cards
    # =========================
    def _create_stat_card(self, parent, title, value, color):
        card = tk.Frame(
            parent,
            bg="#020617",
            bd=0,
            relief="flat",
            highlightthickness=1,
            highlightbackground="#111827",
        )
        card.pack(side="left", expand=True, fill="x", padx=8, pady=10)

        tk.Label(
            card,
            text=title,
            font=("Segoe UI", 10),
            fg="#9CA3AF",
            bg="#020617",
        ).pack(anchor="w", padx=12, pady=(10, 0))

        val_lbl = tk.Label(
            card,
            text=value,
            font=("Segoe UI", 18, "bold"),
            fg=color,
            bg="#020617",
        )
        val_lbl.pack(anchor="w", padx=12, pady=(2, 10))

        return val_lbl

    def refresh_overview_cards(self):
        """Update top statistic cards based on CSV data (global attendance.csv)."""
        attendance_df = get_local_data()
        students_df = (
            pd.read_csv(STUDENTS_FILE) if os.path.exists(STUDENTS_FILE) else None
        )

        total_students = len(students_df) if students_df is not None else 0
        today_str = datetime.now().strftime("%Y-%m-%d")

        total_classes = 0
        present_today = 0
        absent_today = 0

        if attendance_df is not None and not attendance_df.empty:
            df = clean_data(attendance_df)

            total_classes = len(df["Date"].unique())
            present_today = df[df["Date"] == today_str]["USN"].nunique()

            if students_df is not None:
                all_usns = set(students_df["USN"].str.upper())
                present_usns = set(df[df["Date"] == today_str]["USN"].str.upper())
                absent_today = len(all_usns - present_usns)

        self.card_total_students.config(text=str(total_students))
        self.card_present_today.config(text=str(present_today))
        self.card_absent_today.config(text=str(absent_today))
        self.card_total_classes.config(text=str(total_classes))

    # =========================
    # Student registration
    # =========================
    def open_student_registration(self):
        open_registration_window(self.root)

    # =========================
    # Auto-clear scheduler
    # =========================
    def schedule_midnight_check(self):
        self.root.after(60_000, self.check_date_rollover)

    def check_date_rollover(self):
        today = datetime.now().strftime("%Y-%m-%d")
        if today != self.current_date:
            create_backup()
            with open(ATTENDANCE_FILE, "w", encoding="utf-8") as f:
                f.write("USN,Name,Date,Time,Status\n")
            self.current_date = today
        self.schedule_midnight_check()
        self.refresh_overview_cards()

    # =========================
    # Backup Now
    # =========================
    def manual_backup(self):
        path = create_backup()
        if path:
            messagebox.showinfo("Backup Created", f"✔ Backup saved at:\n{path}")
        else:
            messagebox.showerror("Error", "Backup failed.")

    # =========================
    # Export Today
    # =========================
    def export_todays_attendance(self):
        if not os.path.exists(STUDENTS_FILE):
            messagebox.showerror("Error", "students.csv missing!")
            return

        df = get_local_data()
        if df is None:
            messagebox.showinfo("Info", "No attendance data found.")
            return

        df = clean_data(df)
        today = datetime.now().strftime("%Y-%m-%d")

        present_df = df[df["Date"] == today][["USN", "Name", "Time", "Status"]]

        students = pd.read_csv(STUDENTS_FILE)
        students["USN"] = students["USN"].str.upper()

        present_usns = set(present_df["USN"].str.upper())
        all_usns = set(students["USN"])
        absent_usns = all_usns - present_usns

        absent_df = students[students["USN"].isin(absent_usns)][["USN", "Name"]]

        file_path = os.path.join(BASE_DIR, f"Attendance_Today_{today}.xlsx")

        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            present_df.to_excel(writer, sheet_name="Present Today", index=False)
            absent_df.to_excel(writer, sheet_name="Absent Today", index=False)

        messagebox.showinfo("Success", f"✔ Exported today!\n{file_path}")

    # =========================
    # Export Previous Day
    # =========================
    def export_previous_day_attendance(self):
        if not os.path.exists(STUDENTS_FILE):
            messagebox.showerror("Error", "students.csv missing!")
            return

        df = get_local_data()
        if df is None:
            messagebox.showinfo("Info", "No attendance found.")
            return

        prev = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")

        df = clean_data(df)
        present_df = df[df["Date"] == prev][["USN", "Name", "Time", "Status"]]

        if present_df.empty:
            messagebox.showinfo("Info", f"No data for previous day ({prev}).")
            return

        students = pd.read_csv(STUDENTS_FILE)
        students["USN"] = students["USN"].str.upper()

        present_usns = set(present_df["USN"].str.upper())
        all_usns = set(students["USN"])
        absent_usns = all_usns - present_usns

        absent_df = students[students["USN"].isin(absent_usns)][["USN", "Name"]]

        path = os.path.join(BASE_DIR, f"Attendance_PreviousDay_{prev}.xlsx")

        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            present_df.to_excel(writer, sheet_name="Present Prev Day", index=False)
            absent_df.to_excel(writer, sheet_name="Absent Prev Day", index=False)

        messagebox.showinfo("Success", f"✔ Exported previous day!\n{path}")

    # =========================
    # Export Summary
    # =========================
    def export_summary_to_excel(self):
        if not os.path.exists(STUDENTS_FILE) or not os.path.exists(ATTENDANCE_FILE):
            messagebox.showerror("Error", "Required CSV missing!")
            return

        students = pd.read_csv(STUDENTS_FILE)
        df = clean_data(pd.read_csv(ATTENDANCE_FILE))

        total_classes = len(df["Date"].unique())
        summary = []

        for _, st in students.iterrows():
            usn = st["USN"].upper()
            attended = df[df["USN"].str.upper() == usn]["Date"].nunique()
            percent = round(attended / total_classes * 100, 2) if total_classes else 0
            summary.append([usn, st["Name"], attended, total_classes, percent])

        summary_df = pd.DataFrame(
            summary,
            columns=[
                "USN",
                "Name",
                "Classes Attended",
                "Total Classes",
                "Attendance %",
            ],
        )

        today = datetime.now().strftime("%Y-%m-%d")
        path = os.path.join(BASE_DIR, f"Attendance_Summary_Till_{today}.xlsx")
        summary_df.to_excel(path, index=False)

        # Coloring
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill

        wb = load_workbook(path)
        ws = wb.active

        for row in ws.iter_rows(min_row=2):
            percent = float(row[4].value)
            if percent >= 75:
                color = "C6EFCE"
            elif percent >= 50:
                color = "FFF2CC"
            else:
                color = "FFC7CE"
            fill = PatternFill(start_color=color, fill_type="solid")
            for cell in row:
                cell.fill = fill

        wb.save(path)
        wb.close()

        messagebox.showinfo("Success", f"✔ Summary exported!\n{path}")

    # =========================
    # Visualization Menu
    # =========================
    def open_visualization_menu(self):
        df = get_local_data()
        if df is None or df.empty:
            messagebox.showinfo("Info", "No data to visualize.")
            return

        menu = tk.Toplevel(self.root)
        menu.title("📊 Data Visualization")
        menu.geometry("400x220")
        menu.resizable(False, False)
        menu.config(bg="#020617")

        tk.Label(
            menu,
            text="Choose Visualization",
            font=("Segoe UI", 14, "bold"),
            bg="#020617",
            fg="#E5E7EB",
        ).pack(pady=15)

        tk.Button(
            menu,
            text="📊 Class Overview",
            command=self.open_class_overview_window,
            bg="#0F766E",
            fg="white",
            font=("Segoe UI", 11, "bold"),
            relief="flat",
            padx=8,
            pady=6,
        ).pack(pady=8, fill="x", padx=40)

        tk.Button(
            menu,
            text="📈 Student Trend (Select Row)",
            command=self._show_student_trend_internal,
            bg="#7C2D12",
            fg="white",
            font=("Segoe UI", 11, "bold"),
            relief="flat",
            padx=8,
            pady=6,
        ).pack(pady=8, fill="x", padx=40)

    # =========================
    # Class Overview Graph
    # =========================
    def open_class_overview_window(self):
        df = get_local_data()
        if df is None or df.empty:
            return

        df = clean_data(df)

        try:
            from matplotlib.figure import Figure
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
        except Exception:
            messagebox.showerror("Error", "Install matplotlib: pip install matplotlib")
            return

        win = tk.Toplevel(self.root)
        win.title("📊 Class Attendance Overview")
        win.geometry("1000x800")

        fig = Figure(figsize=(9, 8), dpi=100)

        # Trend graph
        trend_ax = fig.add_subplot(211)
        date_counts = df.groupby("Date")["USN"].nunique().reset_index()
        trend_ax.plot(date_counts["Date"], date_counts["USN"], marker="o")
        trend_ax.set_title("Students Present per Day")
        trend_ax.set_xticklabels(date_counts["Date"], rotation=45)

        # Student-wise %
        bar_ax = fig.add_subplot(212)
        students = pd.read_csv(STUDENTS_FILE)
        total = len(df["Date"].unique())

        vals = []
        for _, st in students.iterrows():
            usn = st["USN"].upper()
            present = df[df["USN"].str.upper() == usn]["Date"].nunique()
            percent = round(present / total * 100, 2) if total else 0
            vals.append((st["Name"], percent))

        if vals:
            names, per = zip(*vals)
            bar_ax.bar(names, per)
            bar_ax.set_title("Attendance %")
            bar_ax.set_xticklabels(names, rotation=90)

        canvas = FigureCanvasTkAgg(fig, win)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True)

    # =========================
    # Student Trend
    # =========================
    def on_tree_double_click(self, event):
        self._show_student_trend_internal()

    def _show_student_trend_internal(self):
        df = get_local_data()
        if df is None or df.empty:
            return

        df = clean_data(df)

        selected = self.tree.selection()
        if not selected:
            messagebox.showinfo("Info", "Select a student row first.")
            return

        item = self.tree.item(selected[0])
        values = item.get("values")

        usn = str(values[0]).upper()
        name = str(values[1])

        try:
            from matplotlib.figure import Figure
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
        except Exception:
            messagebox.showerror("Error", "Install matplotlib")
            return

        dates = sorted(df["Date"].astype(str).unique())
        y_vals = []

        for d in dates:
            present = not df[
                (df["USN"].str.upper() == usn)
                & (df["Date"] == d)
                & (df["Status"].str.upper() == "PRESENT")
            ].empty
            y_vals.append(1 if present else 0)

        win = tk.Toplevel(self.root)
        win.title(f"📈 Trend - {name} ({usn})")
        win.geometry("900x600")

        fig = Figure(figsize=(8, 6), dpi=100)
        ax = fig.add_subplot(111)

        ax.plot(dates, y_vals, marker="o")
        ax.set_title(f"Attendance Trend - {name} ({usn})")
        ax.set_ylim(-0.2, 1.2)
        ax.set_yticks([0, 1])
        ax.set_yticklabels(["Absent", "Present"])
        ax.tick_params(axis="x", rotation=45)
        ax.grid(True, linestyle="--", alpha=0.5)

        canvas = FigureCanvasTkAgg(fig, win)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True)

    # =========================
    # Load Google Sheet
    # =========================
    def load_google_data(self):
        df = get_google_sheet_data()
        if df is not None:
            df = clean_data(df)
            self.display_data(df)
            self.refresh_overview_cards()
            messagebox.showinfo("Success", "Loaded Google Sheet!")

    # =========================
    # Load Local CSV
    # =========================
    def load_local_data(self):
        df = get_local_data()
        if df is not None:
            df = clean_data(df)
            self.display_data(df)
            self.refresh_overview_cards()
            messagebox.showinfo("Success", "Loaded local CSV!")

    # =========================
    # Show Today
    # =========================
    def show_today_attendance(self):
        df = get_local_data()
        if df is None:
            return
        today = datetime.now().strftime("%Y-%m-%d")
        df_today = df[df["Date"] == today]
        self.display_data(df_today)
        self.refresh_overview_cards()

    # =========================
    # Clear Attendance
    # =========================
    def clear_previous_attendance(self):
        confirm = messagebox.askyesno(
            "Confirm Clear",
            "⚠ Clear ALL attendance?\nA backup will be created.",
        )
        if not confirm:
            return

        backup_path = create_backup()

        with open(ATTENDANCE_FILE, "w") as f:
            f.write("USN,Name,Date,Time,Status\n")

        for row in self.tree.get_children():
            self.tree.delete(row)

        self.refresh_overview_cards()

        messagebox.showinfo("Success", f"Cleared! Backup saved:\n{backup_path}")

    # =========================
    # Run Face Recognition
    # =========================
    def run_face_recognition_thread(self):
        threading.Thread(target=self.start_face_recognition, daemon=True).start()

    def start_face_recognition(self):
        python_path = os.path.join(BASE_DIR, "venv", "Scripts", "python.exe")
        script_path = os.path.join(BASE_DIR, "main.py")

        subject = self.current_subject.get()
        if not subject:
            subject = "General"

        # Pass subject as an extra CLI argument → main.py can use it
        try:
            subprocess.Popen(
                ["cmd.exe", "/K", python_path, script_path, "auto", subject]
            )
            messagebox.showinfo(
                "Running",
                f"Face recognition started for subject: {subject}!"
            )
        except Exception as e:
            messagebox.showerror(
                "Error",
                f"Could not start face recognition:\n{e}"
            )

    # =========================
    # Export menu (from sidebar)
    # =========================
    def open_export_menu(self):
        menu = tk.Toplevel(self.root)
        menu.title("📤 Export Options")
        menu.geometry("360x230")
        menu.resizable(False, False)
        menu.config(bg="#020617")

        tk.Label(
            menu, text="Choose Export Action",
            font=("Segoe UI", 13, "bold"),
            bg="#020617",
            fg="#E5E7EB",
        ).pack(pady=10)

        tk.Button(
            menu,
            text="📥 Export Today",
            command=self.export_todays_attendance,
            bg="#0EA5E9",
            fg="white",
            font=("Segoe UI", 11, "bold"),
            relief="flat",
            padx=8,
            pady=6,
        ).pack(pady=5, fill="x", padx=40)

        tk.Button(
            menu,
            text="📦 Export Previous Day",
            command=self.export_previous_day_attendance,
            bg="#6B21A8",
            fg="white",
            font=("Segoe UI", 11, "bold"),
            relief="flat",
            padx=8,
            pady=6,
        ).pack(pady=5, fill="x", padx=40)

        tk.Button(
            menu,
            text="📤 Export Summary (Till Today)",
            command=self.export_summary_to_excel,
            bg="#1D4ED8",
            fg="white",
            font=("Segoe UI", 11, "bold"),
            relief="flat",
            padx=8,
            pady=6,
        ).pack(pady=5, fill="x", padx=40)

    # =========================
    # NEW: Search Student (by USN / Name)
    # =========================
    def open_search_student_window(self):
        df = get_local_data()
        if df is None or df.empty:
            messagebox.showinfo("Info", "No attendance data to search.")
            return

        win = tk.Toplevel(self.root)
        win.title("🔍 Search Student")
        win.geometry("380x260")
        win.resizable(False, False)
        win.config(bg="#020617")

        tk.Label(
            win,
            text="Search Student Attendance",
            font=("Segoe UI", 13, "bold"),
            bg="#020617",
            fg="#E5E7EB",
        ).pack(pady=10)

        form = tk.Frame(win, bg="#020617")
        form.pack(pady=5, padx=20, fill="x")

        tk.Label(form, text="USN:", font=("Segoe UI", 10), bg="#020617", fg="#E5E7EB").grid(row=0, column=0, sticky="w")
        usn_entry = tk.Entry(form, font=("Segoe UI", 10))
        usn_entry.grid(row=0, column=1, pady=4, sticky="ew")

        tk.Label(form, text="Name (optional):", font=("Segoe UI", 10), bg="#020617", fg="#E5E7EB").grid(row=1, column=0, sticky="w")
        name_entry = tk.Entry(form, font=("Segoe UI", 10))
        name_entry.grid(row=1, column=1, pady=4, sticky="ew")

        form.columnconfigure(1, weight=1)

        def do_search():
            usn = usn_entry.get().strip().upper()
            name = name_entry.get().strip()

            if not usn and not name:
                messagebox.showwarning("Input", "Enter at least USN or Name.")
                return

            data = get_local_data()
            if data is None or data.empty:
                messagebox.showinfo("Info", "No attendance data found.")
                return

            data = clean_data(data)

            if usn:
                data["USN"] = data["USN"].astype(str).str.upper()
                result = data[data["USN"] == usn]
            else:
                result = data[data["Name"].astype(str).str.lower() == name.lower()]

            if result.empty:
                messagebox.showinfo("No Data", "No attendance found for this student.")
                return

            total_classes = len(data["Date"].unique())
            attended_days = result["Date"].nunique()
            percent = round(attended_days / total_classes * 100, 2) if total_classes else 0

            # try to pick a display name
            if "Name" in result.columns:
                display_name = result.iloc[0]["Name"]
            else:
                display_name = name if name else usn

            msg_lines = [
                f"USN: {usn if usn else 'N/A'}",
                f"Name: {display_name}",
                "",
                f"Total Classes Conducted: {total_classes}",
                f"Days Present: {attended_days}",
                f"Attendance %: {percent:.2f}%",
                "",
                "Dates Present:",
            ]
            present_dates = sorted(result["Date"].unique())
            msg_lines.extend(present_dates)

            messagebox.showinfo("Student Attendance", "\n".join(msg_lines))

        tk.Button(
            win,
            text="Search",
            command=do_search,
            bg="#6366F1",
            fg="white",
            font=("Segoe UI", 10, "bold"),
            relief="flat",
            padx=10,
            pady=5,
        ).pack(pady=15)

    # =========================
    # NEW FEATURE: Subject-wise Absent Students
    # =========================
    def show_absent_subjectwise(self):
        """Show who is absent for the selected subject today."""
        subject = self.current_subject.get()
        if not subject:
            messagebox.showwarning("Subject Missing", "Please select a subject first.")
            return

        # CSV file for this subject (created by main.py)
        subject_file = os.path.join(BASE_DIR, f"Attendance_{subject}.csv")

        if not os.path.exists(STUDENTS_FILE):
            messagebox.showerror("Error", "students.csv missing!")
            return

        students_df = pd.read_csv(STUDENTS_FILE)
        students_df["USN"] = students_df["USN"].str.upper()

        # If NO attendance file exists → ALL are absent
        if not os.path.exists(subject_file):
            absent_df = students_df.copy()
        else:
            att_df = pd.read_csv(subject_file)
            if "USN" not in att_df.columns or "Date" not in att_df.columns:
                messagebox.showerror("Error", f"Invalid format in {subject_file}")
                return

            att_df["USN"] = att_df["USN"].astype(str).str.upper()
            today = datetime.now().strftime("%Y-%m-%d")

            present_today = set(att_df[att_df["Date"] == today]["USN"])
            all_usns = set(students_df["USN"])
            absent_usns = all_usns - present_today
            absent_df = students_df[students_df["USN"].isin(absent_usns)]

        # Show Results in a Popup Window
        win = tk.Toplevel(self.root)
        win.title(f"🚫 Absent Students - {subject}")
        win.geometry("500x450")
        win.resizable(False, False)
        win.config(bg="#020617")

        tk.Label(
            win,
            text=f"Absent Students ({subject}) - Today",
            font=("Segoe UI", 14, "bold"),
            bg="#020617",
            fg="#F97373",
        ).pack(pady=10)

        table_frame = tk.Frame(win, bg="#020617")
        table_frame.pack(fill="both", expand=True, padx=10, pady=10)

        table = ttk.Treeview(
            table_frame,
            columns=("USN", "Name"),
            show="headings",
            height=15
        )
        table.heading("USN", text="USN")
        table.heading("Name", text="Name")
        table.column("USN", width=160, anchor="center")
        table.column("Name", width=260, anchor="w")

        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=table.yview)
        table.configure(yscroll=vsb.set)
        vsb.pack(side="right", fill="y")
        table.pack(fill="both", expand=True)

        for _, row in absent_df.iterrows():
            table.insert("", "end", values=[row["USN"], row["Name"]])

        # Export Button
        def export_absent():
            today = datetime.now().strftime("%Y-%m-%d")
            path = os.path.join(BASE_DIR, f"Absent_{subject}_{today}.xlsx")
            absent_df.to_excel(path, index=False)
            messagebox.showinfo("Exported", f"Absent list saved:\n{path}")

        tk.Button(
            win,
            text="Export to Excel",
            command=export_absent,
            bg="#1D4ED8",
            fg="white",
            font=("Segoe UI", 10, "bold"),
            relief="flat",
            padx=10,
            pady=6
        ).pack(pady=10)

    # =========================
    # Display Data
    # =========================
    def display_data(self, df):
        for row in self.tree.get_children():
            self.tree.delete(row)
        if df is None or df.empty:
            return

        for idx, (_, r) in enumerate(df.iterrows()):
            values = [
                r.get("USN", ""),
                r.get("Name", ""),
                r.get("Date", ""),
                r.get("Time", ""),
                r.get("Status", ""),
            ]
            tag = "evenrow" if idx % 2 == 0 else "oddrow"
            self.tree.insert("", "end", values=values, tags=(tag,))


# -----------------------------
# 🚀 Run App
# -----------------------------
if __name__ == "__main__":
    root = tk.Tk()
    app = AttendanceDashboard(root)
    root.mainloop()
